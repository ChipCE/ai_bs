import openpyxl
from datetime import datetime, date
from calendar import monthrange
import uuid
import os
import time  # for filesystem flush on some platforms
import re
import unicodedata
import shutil  # for backup copy
import zipfile
import msvcrt

# --------------------------------------------------------------------------- #
# Date helpers for multi-month support                                       #
# --------------------------------------------------------------------------- #

def _iter_month_ranges(start_date: date, end_date: date):
    """
    Yield (m_start, m_end) pairs where each pair covers the portion of the
    original [start_date, end_date] that falls inside a single month.

    The first span starts at `start_date`, the last span ends at `end_date`.
    """
    cur_year, cur_month = start_date.year, start_date.month

    while True:
        last_day = monthrange(cur_year, cur_month)[1]
        span_start = start_date if (cur_year == start_date.year and cur_month == start_date.month) \
            else date(cur_year, cur_month, 1)

        # decide span_end
        if cur_year == end_date.year and cur_month == end_date.month:
            span_end = end_date
        else:
            span_end = date(cur_year, cur_month, last_day)

        yield span_start, span_end

        # break if we just yielded the final month
        if cur_year == end_date.year and cur_month == end_date.month:
            break

        # advance month
        if cur_month == 12:
            cur_year += 1
            cur_month = 1
        else:
            cur_month += 1


def _get_month_sheet_name(target_date):
    """Get the sheet name for a given date in 'yy年M月' format."""
    year_short = str(target_date.year)[-2:]  # Get last 2 digits of year
    month = target_date.month
    return f"{year_short}年{month}月"


def _find_device_row(sheet, device_name):
    """Find the row index for a device name in column B."""
    for row in range(9, sheet.max_row + 1):
        if sheet.cell(row=row, column=2).value == device_name:
            return row
    return None


def _get_date_column(sheet, day):
    """Find the column index for a specific day, tolerant to header formats."""
    # Primary: row 8
    for col in range(3, sheet.max_column + 1):
        if _normalize_header_day(sheet.cell(row=8, column=col).value) == day:
            return col
    # Fallback: nearby rows in case layout shifted
    for hdr_row in (7, 9, 10):
        if hdr_row <= sheet.max_row:
            for col in range(3, sheet.max_column + 1):
                if _normalize_header_day(sheet.cell(row=hdr_row, column=col).value) == day:
                    return col
    return None


def _get_date_columns(sheet, start_date, end_date):
    """Get list of column indices for a date range using tolerant header parsing."""
    columns = []
    for d in range(start_date.day, end_date.day + 1):
        col = _get_date_column(sheet, d)
        if col is None:
            raise ValueError(f"日付が見つかりません: {d}")
        columns.append(col)
    return columns


def _generate_booking_id():
    """Generate a unique booking ID."""
    return str(uuid.uuid4())[:8]

# --------------------------------------------------------------------------- #
# Simple file-lock and safe save helper                                       #
# --------------------------------------------------------------------------- #


class _FileLock:
    """Directory-based lock to guard workbook writes (best-effort, cross-platform)."""

    def __init__(self, target_path, timeout: float = 10.0, interval: float = 0.1):
        self.lock_dir = target_path + ".lock"
        self.target_path = target_path
        self.timeout = timeout
        self.interval = interval

    def __enter__(self):
        start = time.time()
        while True:
            try:
                os.mkdir(self.lock_dir)
                break
            except FileExistsError:
                if time.time() - start > self.timeout:
                    raise TimeoutError("ファイルが使用中です。しばらくしてから再度お試しください。")
                time.sleep(self.interval)
        # Wait while Excel (or another proc) keeps file locked
        start = time.time()
        while _is_file_in_use(self.target_path):
            if time.time() - start > self.timeout:
                raise TimeoutError("Excelでファイルが開かれています。閉じてから再度お試しください。")
            time.sleep(self.interval)
        return self

    def __exit__(self, exc_type, exc, tb):
        try:
            os.rmdir(self.lock_dir)
        except Exception:
            pass


def _safe_save_workbook(wb, excel_path):
    """
    Save workbook atomically and keep a timestamped backup of the original file.

    Returns
    -------
    (bak_path, pre_size, post_size)
    """
    base_dir = os.path.dirname(excel_path)
    tmp_path = os.path.join(base_dir, f"._tmp_{uuid.uuid4().hex}.xlsx")
    bak_path = None

    pre_size = os.path.getsize(excel_path) if os.path.exists(excel_path) else None

    if os.path.exists(excel_path):
        ts = datetime.now().strftime("%Y%m%d%H%M%S")
        bak_path = os.path.join(base_dir, f"{os.path.basename(excel_path)}.{ts}.bak")
        try:
            shutil.copy2(excel_path, bak_path)
        except Exception:
            bak_path = None  # backup failed, continue anyway

    wb.save(tmp_path)
    # ensure bytes are flushed to disk
    with open(tmp_path, "rb") as _fh:
        try:
            os.fsync(_fh.fileno())
        except OSError:
            pass
    # close workbook before replacing to release file handles on Windows
    try:
        wb.close()
    except Exception:
        pass
    os.replace(tmp_path, excel_path)

    post_size = os.path.getsize(excel_path)
    return bak_path, pre_size, post_size


def _validate_or_restore(
    excel_path,
    bak_path,
    pre_size,
    post_size,
    expected_sheets,
    max_ratio_diff: float,
    min_size_bytes: int = 4096,
):
    """
    Validate that workbook was saved correctly; restore backup on suspicion.

    Returns True if validation passed, False if restore executed.
    """
    ok = True
    # -------- ZIP structure check -------- #
    try:
        with zipfile.ZipFile(excel_path, "r") as zf:
            required = {"[Content_Types].xml", "xl/workbook.xml"}
            names = set(zf.namelist())
            if not required.issubset(names) or zf.testzip() is not None:
                ok = False
    except Exception:
        ok = False

    if not ok:
        if bak_path and os.path.exists(bak_path):
            try:
                shutil.copy2(bak_path, excel_path)
            except Exception:
                pass
        return False
    try:
        wb_v = openpyxl.load_workbook(excel_path)
        try:
            # sheet presence
            for s in expected_sheets:
                if s not in wb_v.sheetnames:
                    ok = False
                    break
        finally:
            wb_v.close()
    except Exception:
        ok = False

    # size checks
    if ok and pre_size is not None and post_size is not None:
        try:
            ratio = abs(post_size - pre_size) / max(pre_size, 1)
            if ratio > max_ratio_diff:
                ok = False
        except Exception:
            ok = False
    if ok and post_size is not None and post_size < min_size_bytes:
        ok = False

    if ok:
        return True

    # attempt ZIP validation failed, or other checks failed
    # attempt restore
    if bak_path and os.path.exists(bak_path):
        try:
            shutil.copy2(bak_path, excel_path)
        except Exception:
            pass
    return False


# --------------------------------------------------------------------------- #
# File lock helper                                                            #
# --------------------------------------------------------------------------- #

def _is_file_in_use(path):
    """Best-effort check on Windows: returns True if the file is locked."""
    if not os.path.exists(path):
        return False
    try:
        with open(path, "r+b") as fh:
            try:
                msvcrt.locking(fh.fileno(), msvcrt.LK_NBLCK, 1)
                msvcrt.locking(fh.fileno(), msvcrt.LK_UNLCK, 1)
                return False
            except OSError:
                return True
    except OSError:
        return True


# --------------------------------------------------------------------------- #
# Config                                                                      #
# --------------------------------------------------------------------------- #

_RATIO_LIMIT = float(os.getenv("EXCEL_SIZE_DIFF_RATIO", "0.5"))
_WRITE_MODE = os.getenv("EXCEL_WRITE_MODE", "").strip().lower()


# --------------------------------------------------------------------------- #
# Helpers for booking log sheet                                              #
# --------------------------------------------------------------------------- #

def _ensure_booking_log_sheet(wb):
    """
    Ensure that a sheet named '予約ログ' exists and contains the correct headers.

    This is called before writing to the log so that first-time or
    手動で削除されたワークブックでも自動で復旧できる。
    """
    if '予約ログ' not in wb.sheetnames:
        sheet = wb.create_sheet('予約ログ')
        headers = [
            '予約ID', '予約日時', '予約者名', '内線番号', '職番',
            'デモ機名', '予約開始日', '予約終了日', 'ステータス'
        ]
        for i, h in enumerate(headers, 1):
            sheet.cell(row=1, column=i, value=h)


# --------------------------------------------------------------------------- #
# Utility iterator to enumerate device rows                                  #
# --------------------------------------------------------------------------- #

def _iter_device_rows(sheet):
    """
    Yield (row_index, device_name) tuples for each device registered
    in column B starting from row 9.
    """
    for row in range(9, sheet.max_row + 1):
        name = sheet.cell(row=row, column=2).value
        if name:
            yield row, str(name)


def _normalize_header_day(value):
    """
    Convert various header cell values to an int day-of-month.

    Acceptable formats:
      * int / float  -> 1..31
      * datetime / date -> .day
      * str -> '1', '01', '1日', '１' etc. (full-width digits ok)

    Returns:
        int | None
    """
    from datetime import datetime as _dt
    from datetime import date as _date

    if isinstance(value, (int, float)):
        try:
            return int(value)
        except Exception:
            return None
    if isinstance(value, (_dt, _date)):
        return int(value.day)
    if isinstance(value, str):
        s = unicodedata.normalize("NFKC", value).strip()
        m = re.search(r"(\d{1,2})", s)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                return None
    return None


# --------------------------------------------------------------------------- #
# COM-based helpers                                                          #
# --------------------------------------------------------------------------- #

try:
    import win32com.client as win32  # type: ignore
except Exception:
    win32 = None
try:
    import pythoncom  # type: ignore
except Exception:
    pythoncom = None

def _com_ensure_booking_log_sheet(wb):
    try:
        wb.Worksheets("予約ログ")
    except Exception:
        ws = wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count))
        ws.Name = "予約ログ"
        headers = [
            '予約ID','予約日時','予約者名','内線番号','職番',
            'デモ機名','予約開始日','予約終了日','ステータス'
        ]
        for i, h in enumerate(headers, 1):
            ws.Cells(1, i).Value = h


def _com_find_device_row(ws, device_name):
    last_row = ws.Cells(ws.Rows.Count, 2).End(-4162).Row  # xlUp
    for r in range(9, last_row + 1):
        if str(ws.Cells(r, 2).Value).strip() == str(device_name):
            return r
    return None


def _com_get_date_columns(ws, m_start, m_end):
    cols = []
    last_col = ws.Cells(8, ws.Columns.Count).End(-4159).Column  # xlToLeft
    for c in range(3, last_col + 1):
        val = ws.Cells(8, c).Value
        day = _normalize_header_day(val)
        if day is not None and m_start.day <= day <= m_end.day:
            cols.append(c)
    if not cols:
        raise ValueError("日付が見つかりません")
    return cols


def _com_book(excel_path, device_name, start_date, end_date, user_info):
    if win32 is None:
        raise RuntimeError("win32com is unavailable; cannot use COM write mode")
    if pythoncom is not None:
        pythoncom.CoInitialize()
    excel = win32.Dispatch("Excel.Application")
    excel.DisplayAlerts = False
    excel.Visible = False
    try:
        wb = excel.Workbooks.Open(excel_path, UpdateLinks=0, ReadOnly=False)
        try:
            booking_id = _generate_booking_id()
            marker = f"C:{booking_id}"
            for m_start, m_end in _iter_month_ranges(start_date, end_date):
                sheet_name = _get_month_sheet_name(m_start)
                try:
                    ws = wb.Worksheets(sheet_name)
                except Exception:
                    raise ValueError(f"シートが見つかりません: {sheet_name}")
                row = _com_find_device_row(ws, device_name)
                if row is None:
                    raise ValueError(f"デモ機が見つかりません: {device_name}")
                for c in _com_get_date_columns(ws, m_start, m_end):
                    ws.Cells(row, c).Value = marker

            _com_ensure_booking_log_sheet(wb)
            log = wb.Worksheets("予約ログ")
            last = log.Cells(log.Rows.Count, 1).End(-4162).Row  # xlUp
            next_row = last + 1 if last >= 1 else 2
            now = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
            log.Cells(next_row, 1).Value = booking_id
            log.Cells(next_row, 2).Value = now
            log.Cells(next_row, 3).Value = user_info.get('name', '')
            log.Cells(next_row, 4).Value = user_info.get('extension', '')
            log.Cells(next_row, 5).Value = user_info.get('employee_id', '')
            log.Cells(next_row, 6).Value = device_name
            log.Cells(next_row, 7).Value = start_date.strftime("%Y-%m-%d")
            log.Cells(next_row, 8).Value = end_date.strftime("%Y-%m-%d")
            log.Cells(next_row, 9).Value = '予約中'

            # Set font color to black for the new log row
            for i in range(1, 10):
                log.Cells(next_row, i).Font.ColorIndex = 1

            wb.Save()
            return booking_id
        finally:
            wb.Close(SaveChanges=True)
    finally:
        excel.Quit()
        if pythoncom is not None:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass


def _com_cancel(excel_path, booking_id):
    if win32 is None:
        raise RuntimeError("win32com is unavailable; cannot use COM write mode")
    if pythoncom is not None:
        pythoncom.CoInitialize()
    excel = win32.Dispatch("Excel.Application")
    excel.DisplayAlerts = False
    excel.Visible = False
    try:
        wb = excel.Workbooks.Open(excel_path, UpdateLinks=0, ReadOnly=False)
        try:
            log = wb.Worksheets("予約ログ")
            last = log.Cells(log.Rows.Count, 1).End(-4162).Row
            target_row = None
            for r in range(2, last + 1):
                if str(log.Cells(r, 1).Value).strip() == str(booking_id):
                    target_row = r
                    break
            if target_row is None:
                raise ValueError(f"予約IDが見つかりません: {booking_id}")
            device_name = str(log.Cells(target_row, 6).Value).strip()
            start_date = datetime.strptime(str(log.Cells(target_row, 7).Value).strip(), "%Y-%m-%d").date()
            end_date = datetime.strptime(str(log.Cells(target_row, 8).Value).strip(), "%Y-%m-%d").date()

            for m_start, m_end in _iter_month_ranges(start_date, end_date):
                sheet_name = _get_month_sheet_name(m_start)
                try:
                    ws = wb.Worksheets(sheet_name)
                except Exception:
                    raise ValueError(f"シートが見つかりません: {sheet_name}")
                row = _com_find_device_row(ws, device_name)
                if row is None:
                    raise ValueError(f"デモ機が見つかりません: {device_name}")
                prefix = f"C:{booking_id}"
                last_col = ws.Cells(8, ws.Columns.Count).End(-4159).Column
                for c in range(3, last_col + 1):
                    val = ws.Cells(8, c).Value
                    day = _normalize_header_day(val)
                    if day is not None and m_start.day <= day <= m_end.day:
                        cell_val = ws.Cells(row, c).Value
                        if isinstance(cell_val, str) and cell_val.startswith(prefix):
                            ws.Cells(row, c).Value = None

            wb.Save()
        finally:
            wb.Close(SaveChanges=True)
    finally:
        excel.Quit()
        if pythoncom is not None:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass


def check_availability(excel_path, device_name, start_date, end_date):
    """Check if a device is available for the specified date range."""
    wb = openpyxl.load_workbook(excel_path)
    try:
        for m_start, m_end in _iter_month_ranges(start_date, end_date):
            sheet_name = _get_month_sheet_name(m_start)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"シートが見つかりません: {sheet_name}")

            sheet = wb[sheet_name]
            device_row = _find_device_row(sheet, device_name)
            if device_row is None:
                raise ValueError(f"デモ機が見つかりません: {device_name}")

            for col in _get_date_columns(sheet, m_start, m_end):
                if sheet.cell(row=device_row, column=col).value is not None:
                    return False
        return True
    finally:
        wb.close()


def find_available_device(excel_path, device_type, start_date, end_date):
    """
    Find the first available device of a given *type* for the specified period.

    Args:
        excel_path (str): Path to workbook
        device_type (str): Prefix of the device name, e.g., \"FE\" or \"PC\"
        start_date (date): Start date (inclusive)
        end_date (date): End date (inclusive)

    Returns:
        str | None: The first available device name or None if none found.
    """
    wb = openpyxl.load_workbook(excel_path)
    try:
        # collect candidate names from the start-month sheet
        start_sheet_name = _get_month_sheet_name(start_date)
        if start_sheet_name not in wb.sheetnames:
            raise ValueError(f"シートが見つかりません: {start_sheet_name}")

        start_sheet = wb[start_sheet_name]
        candidates = [
            name for _row, name in _iter_device_rows(start_sheet)
            if str(name).startswith(device_type)
        ]

        for dev_name in candidates:
            ok = True
            for m_start, m_end in _iter_month_ranges(start_date, end_date):
                sheet_name = _get_month_sheet_name(m_start)
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"シートが見つかりません: {sheet_name}")
                sheet = wb[sheet_name]
                row = _find_device_row(sheet, dev_name)
                if row is None:
                    raise ValueError(f"デモ機が見つかりません: {dev_name}")
                cols = _get_date_columns(sheet, m_start, m_end)
                if any(sheet.cell(row=row, column=c).value is not None for c in cols):
                    ok = False
                    break
            if ok:
                return dev_name
        return None
    finally:
        wb.close()


def book(excel_path, device_name, start_date, end_date, user_info):
    """
    Book a device for the specified date range.
    
    Args:
        excel_path: Path to the Excel file
        device_name: Name of the device to book
        start_date: Start date (inclusive)
        end_date: End date (inclusive)
        user_info: Dictionary with user information (name, extension, employee_id)
        
    Returns:
        str: Booking ID
        
    Raises:
        ValueError: If the device is already booked for any day in the range
    """
    if _WRITE_MODE == 'com':
        # availability check still uses openpyxl (read-only), but it only reads
        booking_id = _com_book(excel_path, device_name, start_date, end_date, user_info)
        return booking_id
    
    # First check if available
    if not check_availability(excel_path, device_name, start_date, end_date):
        raise ValueError(f"指定された期間にデモ機 '{device_name}' は既に予約されています")

    with _FileLock(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        try:
            # Mark cells across all months
            booking_id = _generate_booking_id()
            booking_marker = f"C:{booking_id}"
            for m_start, m_end in _iter_month_ranges(start_date, end_date):
                sheet_name = _get_month_sheet_name(m_start)
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"シートが見つかりません: {sheet_name}")
                sheet_m = wb[sheet_name]
                device_row_m = _find_device_row(sheet_m, device_name)
                if device_row_m is None:
                    raise ValueError(f"デモ機が見つかりません: {device_name}")
                for col in _get_date_columns(sheet_m, m_start, m_end):
                    sheet_m.cell(row=device_row_m, column=col, value=booking_marker)

            # Ensure booking log sheet exists, then add entry
            _ensure_booking_log_sheet(wb)
            log_sheet = wb['予約ログ']
            next_row = log_sheet.max_row + 1
            now = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
            start_date_str = start_date.strftime("%Y-%m-%d")
            end_date_str = end_date.strftime("%Y-%m-%d")
            log_sheet.cell(row=next_row, column=1, value=booking_id)
            log_sheet.cell(row=next_row, column=2, value=now)
            log_sheet.cell(row=next_row, column=3, value=user_info.get('name', ''))
            log_sheet.cell(row=next_row, column=4, value=user_info.get('extension', ''))
            log_sheet.cell(row=next_row, column=5, value=user_info.get('employee_id', ''))
            log_sheet.cell(row=next_row, column=6, value=device_name)
            log_sheet.cell(row=next_row, column=7, value=start_date_str)
            log_sheet.cell(row=next_row, column=8, value=end_date_str)
            log_sheet.cell(row=next_row, column=9, value='予約中')

            # --- atomic save & validate ---------------------------------- #
            bak_path, pre_size, post_size = _safe_save_workbook(wb, excel_path)
            # expected sheets: all months touched + 予約ログ
            expected = {"予約ログ"}
            for m_start, _m_end in _iter_month_ranges(start_date, end_date):
                expected.add(_get_month_sheet_name(m_start))
            ok = _validate_or_restore(
                excel_path,
                bak_path,
                pre_size,
                post_size,
                list(expected),
                max_ratio_diff=_RATIO_LIMIT,
            )
            if not ok:
                raise IOError("ファイル保存に失敗しました。バックアップから復旧しました。")
            return booking_id
        finally:
            wb.close()


# --------------------------------------------------------------------------- #
# Public helper: list cancellable bookings                                   #
# --------------------------------------------------------------------------- #

def list_cancellable_bookings(excel_path, user_info):
    """
    Return a list of active bookings ('予約中') for the given user.

    Matching keys: name, extension, employee_id.
    Each item is a dict::

        {
            'booking_id': str,
            'device_name': str,
            'start_date': 'YYYY-MM-DD',
            'end_date':   'YYYY-MM-DD',
        }
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    try:
        try:
            log = wb['予約ログ']
        except KeyError:
            return []

        results = []
        name = str(user_info.get('name', '') or '')
        ext = str(user_info.get('extension', '') or '')
        emp = str(user_info.get('employee_id', '') or '')

        for row in range(2, log.max_row + 1):
            status = str(log.cell(row=row, column=9).value or '').strip()
            if status != '予約中':
                continue

            # user matching
            uname = str(log.cell(row=row, column=3).value or '').strip()
            uext = str(log.cell(row=row, column=4).value or '').strip()
            uemp = str(log.cell(row=row, column=5).value or '').strip()

            if not (
                (name and uname == name)
                or (ext and uext == ext)
                or (emp and uemp == emp)
            ):
                continue

            rid = str(log.cell(row=row, column=1).value or '').strip()
            dev = str(log.cell(row=row, column=6).value or '').strip()
            sd = str(log.cell(row=row, column=7).value or '').strip()
            ed = str(log.cell(row=row, column=8).value or '').strip()
            results.append(
                {
                    'booking_id': rid,
                    'device_name': dev,
                    'start_date': sd,
                    'end_date': ed,
                }
            )
        return results
    finally:
        wb.close()


def cancel(excel_path, booking_id):
    """
    Cancel a booking by ID.
    
    Args:
        excel_path: Path to the Excel file
        booking_id: ID of the booking to cancel
        
    Raises:
        ValueError: If the booking ID is not found
    """
    if _WRITE_MODE == 'com':
        _com_cancel(excel_path, booking_id)
        return
        
    with _FileLock(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        try:
            log_sheet = wb['予約ログ']
            booking_row = None
            for row in range(2, log_sheet.max_row + 1):
                if log_sheet.cell(row=row, column=1).value == booking_id:
                    booking_row = row
                    break
            if booking_row is None:
                raise ValueError(f"予約IDが見つかりません: {booking_id}")

            device_name = log_sheet.cell(row=booking_row, column=6).value
            start_date = datetime.strptime(log_sheet.cell(row=booking_row, column=7).value, "%Y-%m-%d").date()
            end_date = datetime.strptime(log_sheet.cell(row=booking_row, column=8).value, "%Y-%m-%d").date()

            for m_start, m_end in _iter_month_ranges(start_date, end_date):
                sheet_name = _get_month_sheet_name(m_start)
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"シートが見つかりません: {sheet_name}")
                sheet = wb[sheet_name]
                device_row = _find_device_row(sheet, device_name)
                if device_row is None:
                    raise ValueError(f"デモ機が見つかりません: {device_name}")
                marker_prefix = f"C:{booking_id}"
                for col in range(3, sheet.max_column + 1):
                    header_val = sheet.cell(row=8, column=col).value
                    day = _normalize_header_day(header_val)
                    if day is not None and m_start.day <= day <= m_end.day:
                        cell_val = sheet.cell(row=device_row, column=col).value
                        if isinstance(cell_val, str) and cell_val.startswith(marker_prefix):
                            sheet.cell(row=device_row, column=col, value=None)

            # --- atomic save & validate (same logic as book) -------------- #
            bak_path, pre_size, post_size = _safe_save_workbook(wb, excel_path)
            expected = {"予約ログ"}
            for m_start, _m_end in _iter_month_ranges(start_date, end_date):
                expected.add(_get_month_sheet_name(m_start))
            ok = _validate_or_restore(
                excel_path,
                bak_path,
                pre_size,
                post_size,
                list(expected),
                max_ratio_diff=_RATIO_LIMIT,
            )
            if not ok:
                raise IOError("ファイル保存に失敗しました。バックアップから復旧しました。")
        finally:
            wb.close()
